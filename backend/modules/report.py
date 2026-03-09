# backend/modules/report.py
import pandas as pd
import io
import re
import datetime
import numpy as np
from scipy.interpolate import make_interp_spline
import matplotlib
matplotlib.use('Agg') 
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from telebot.types import InlineKeyboardMarkup, InlineKeyboardButton
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from backend.database.repository import DatabaseRepo
from backend.utils.formatter import format_currency

class ReportModule:
    def __init__(self):
        self.db = DatabaseRepo()

    def _process_data(self):
        data = self.db.get_report_raw_data()
        crypto_rate = float(data['settings'].get('crypto_rate', 25000))
        
        total_assets = 0
        total_in = 0
        total_out = 0
        wallet_stats = {}
        for w in data['wallets']:
            wid = w['id']
            if wid == 'CASH':
                total_in = w['total_in'] or 0
                total_out = w['total_out'] or 0
                total_assets += w['balance'] or 0
            wallet_stats[wid] = {
                'balance': w['balance'] or 0, 
                'assets': 0, 
                'realized': 0, 
                'unrealized': 0,
                'total_in': w['total_in'] or 0,
                'total_out': w['total_out'] or 0
            }

        symbol_stats = {}
        for h in data['holdings']:
            wid = h['wallet_id']
            sym = h['symbol']
            qty = h['quantity']
            c_price = h['current_price']
            cost = h['cost_basis_vnd']
            
            cur_val = (qty * c_price * crypto_rate) if wid == 'CRYPTO' else (qty * c_price)
            unrealized = cur_val - cost
            
            wallet_stats[wid]['assets'] += cur_val
            wallet_stats[wid]['unrealized'] += unrealized
            total_assets += cur_val
            
            if sym not in symbol_stats:
                symbol_stats[sym] = {'wallet': wid, 'total_pl': 0, 'val': cur_val}
            symbol_stats[sym]['total_pl'] += unrealized

        wins = 0
        losses = 0
        total_realized = 0
        
        current_cap = 0
        current_realized_pl = 0
        peak_nav_proxy = 0

        transactions = sorted(data['transactions'], key=lambda x: x['id'])
        for t in transactions:
            pl = t['realized_pl']
            if pl is not None:
                total_realized += pl
                wallet_stats[t['wallet_id']]['realized'] += pl
                sym = t['symbol']
                if sym:
                    if sym not in symbol_stats: symbol_stats[sym] = {'wallet': t['wallet_id'], 'total_pl': 0, 'val': 0}
                    symbol_stats[sym]['total_pl'] += pl
                
            if t['type'] == 'BAN' and pl is not None:
                if pl > 0: wins += 1
                elif pl < 0: losses += 1
            
            if t['type'] in ['NAP', 'RUT'] and t['wallet_id'] == 'CASH':
                current_cap += (t['amount'] or 0)
            if pl is not None:
                current_realized_pl += pl
            nav_proxy = current_cap + current_realized_pl
            if nav_proxy > peak_nav_proxy:
                peak_nav_proxy = nav_proxy

        total_trades = wins + losses
        win_rate = (wins / total_trades * 100) if total_trades > 0 else 0
        
        peak_nav = max(peak_nav_proxy, total_assets)
        max_drawdown = ((peak_nav - total_assets) / peak_nav * 100) if peak_nav > 0 else 0
        
        sorted_symbols = sorted(symbol_stats.items(), key=lambda x: x[1]['total_pl'], reverse=True)
        top_winners = [s for s in sorted_symbols if s[1]['total_pl'] > 0][:2]
        top_losers = [s for s in reversed(sorted_symbols) if s[1]['total_pl'] < 0][:2]

        return {
            "total_assets": total_assets,
            "total_in": total_in,
            "total_out": total_out,
            "net_cashflow": total_in - total_out,
            "total_realized": total_realized,
            "total_unrealized": sum(w['unrealized'] for w in wallet_stats.values()),
            "total_pl": total_assets - (total_in - total_out),
            "win_rate": win_rate,
            "max_drawdown": max_drawdown,
            "wins": wins,
            "losses": losses,
            "wallets": wallet_stats,
            "top_winners": top_winners,
            "top_losers": top_losers,
            "raw_data": data
        }

    def _generate_nav_chart_io(self, stats, raw):
        capital_history = []; date_history = []; current_calc_capital = 0
        transactions = sorted(raw['transactions'], key=lambda x: x['id'])
        for t in transactions:
            if t['type'] in ['NAP', 'RUT'] and t['wallet_id'] == 'CASH':
                current_calc_capital += (t['amount'] or 0)
                capital_history.append(current_calc_capital)
                tx_date = datetime.date.today()
                if t.get('note'):
                    match = re.search(r'\[(\d{4}-\d{2}-\d{2})\]', str(t['note']))
                    if match:
                        tx_date = datetime.datetime.strptime(match.group(1), '%Y-%m-%d').date()
                date_history.append(tx_date)

        if not capital_history:
            capital_history = [stats['net_cashflow']]; date_history = [datetime.date.today()]

        diff = stats['net_cashflow'] - capital_history[-1]
        capital_history = [val + diff for val in capital_history]

        if len(capital_history) == 1:
            capital_history.insert(0, capital_history[0]); date_history.insert(0, date_history[0] - datetime.timedelta(days=30))

        daily_data = {d: c for d, c in zip(date_history, capital_history)}
        u_dates = sorted(daily_data.keys()); u_caps = [daily_data[d] for d in u_dates]

        fig, ax = plt.subplots(figsize=(10, 5))
        current_assets = stats['total_assets']
        
        if len(u_dates) >= 4:
            x_num = mdates.date2num(u_dates); x_smooth = np.linspace(x_num.min(), x_num.max(), 300)
            spl = make_interp_spline(x_num, u_caps, k=3)
            y_smooth = np.clip(spl(x_smooth), 0, None) 
            ax.fill_between(x_smooth, y_smooth, color='#1f77b4', alpha=0.15)
            ax.plot(x_smooth, y_smooth, linestyle='-', color='#1f77b4', label='Vốn Nạp Ròng', linewidth=2.5)
            ax.plot(x_num, u_caps, 'o', color='#1f77b4', markersize=4, alpha=0.5)
            last_date_num = x_num[-1]
            ax.plot([last_date_num, last_date_num], [u_caps[-1], current_assets], color='#d62728', linestyle='--', linewidth=2)
            ax.plot(last_date_num, current_assets, marker='o', color='#d62728', markersize=8, label='Tài Sản (NAV)')
        else:
            ax.fill_between(u_dates, u_caps, color='#1f77b4', alpha=0.15)
            ax.plot(u_dates, u_caps, marker='o', linestyle='-', color='#1f77b4', label='Vốn Nạp Ròng', linewidth=2.5)
            last_date = u_dates[-1]
            ax.plot([last_date, last_date], [u_caps[-1], current_assets], color='#d62728', linestyle='--', linewidth=2)
            ax.plot(last_date, current_assets, marker='o', color='#d62728', markersize=8, label='Tài Sản (NAV)')

        ax.xaxis.set_major_formatter(mdates.DateFormatter('%m/%Y'))
        plt.title('BIỂU ĐỒ BIẾN ĐỘNG VỐN & TÀI SẢN (NAV)', fontsize=13, fontweight='bold', pad=15)
        plt.grid(True, linestyle='--', alpha=0.6); plt.legend(loc='upper left')
        
        def format_func(value, tick_number):
            if abs(value) >= 1_000_000_000: return f"{value / 1_000_000_000:.1f} Tỷ"
            elif abs(value) >= 1_000_000: return f"{value / 1_000_000:.0f} Tr"
            return f"{value:,.0f}"
        ax.yaxis.set_major_formatter(plt.FuncFormatter(format_func))
        plt.tight_layout()
        buf = io.BytesIO(); plt.savefig(buf, format='png', dpi=120); plt.close(); buf.seek(0)
        return buf

    def generate_chart_bytes(self):
        stats = self._process_data()
        return self._generate_nav_chart_io(stats, stats['raw_data']).getvalue()

    def get_telegram_report(self):
        stats = self._process_data()
        pl_icon = "🟢" if stats['total_pl'] >= 0 else "🔴"
        sign = "+" if stats['total_pl'] > 0 else ""
        nav_roi = (stats['total_pl'] / stats['net_cashflow'] * 100) if stats['net_cashflow'] > 0 else 0
        goal_text = stats['raw_data']['settings'].get('goal', 'lai 10%')
        goal_pct = 10
        try:
            m = re.search(r'\d+', goal_text)
            if m: goal_pct = float(m.group())
        except: pass
        goal_progress = (nav_roi / goal_pct * 100) if goal_pct > 0 else 0
        
        msg = f"📊 **BÁO CÁO QUẢN TRỊ DANH MỤC**\n━━━━━━━━━━━━━━━━━━━\n"
        msg += f"🎯 **HIỆU QUẢ ĐẦU TƯ (NAV)**\n"
        msg += f"💰 Tổng tài sản: {format_currency(stats['total_assets'])}\n"
        msg += f"📈 Lãi/Lỗ tổng: {sign}{format_currency(stats['total_pl'])} ({pl_icon} {sign}{nav_roi:.1f}%)\n"
        msg += f"🏁 Tiến độ mục tiêu: Đạt {goal_progress:.1f}% ({goal_text})\n"
        msg += f"⚖️ Win Rate: {stats['win_rate']:.1f}% ({stats['wins']} Lãi / {stats['losses']} Lỗ)\n"
        msg += f"📉 Max Drawdown: -{stats['max_drawdown']:.1f}%\n"
        msg += f"────────────\n⚖️ **PHÂN BỔ TỶ TRỌNG & HIỆU SUẤT**\n"
        for wid in ['STOCK', 'CRYPTO']:
            w_assets = stats['wallets'][wid]['assets'] + stats['wallets'][wid]['balance']
            pct = (w_assets / stats['total_assets'] * 100) if stats['total_assets'] > 0 else 0
            w_pl = stats['wallets'][wid]['realized'] + stats['wallets'][wid]['unrealized']
            w_icon = "🟢" if w_pl >= 0 else "🔴"
            msg += f"• {wid}: {format_currency(w_assets)} ({pct:.1f}%) | Hiệu suất: {w_icon}\n"
        msg += f"────────────\n🔍 **LỢI NHUẬN ĐẾN TỪ ĐÂU?**\n"
        msg += f"💵 Lãi đã chốt: {format_currency(stats['total_realized'])}\n📄 Lãi chưa chốt: {format_currency(stats['total_unrealized'])}\n"
        if stats['top_losers']:
            msg += f"────────────\n⚠️ **TÀI SẢN KÉO LÙI DANH MỤC**\n"
            for sym, data in stats['top_losers']:
                msg += f"• {sym} ({data['wallet']}): 🔴 {format_currency(data['total_pl'])}\n"
        msg += f"━━━━━━━━━━━━━━━━━━━"
        markup = InlineKeyboardMarkup(row_width=1)
        markup.add(InlineKeyboardButton("📈 Biểu đồ NAV", callback_data="view_nav_chart"),
                   InlineKeyboardButton("📊 Xuất Excel Full", callback_data="export_excel_report"))
        return msg, markup

    def generate_excel_bytes(self):
        stats = self._process_data(); raw = stats['raw_data']
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 1. Dashboard, Portfolio, Performance
            pd.DataFrame({"Chỉ Số": ["Tổng Tài Sản", "Tổng Nạp", "Tổng Rút", "Tiền Mặt", "Lãi/Lỗ Tổng", "Win Rate (%)", "Max Drawdown (%)"], 
                          "Giá Trị": [stats['total_assets'], stats['total_in'], stats['total_out'], stats['wallets']['CASH']['balance'], stats['total_pl'], stats['win_rate'], -stats['max_drawdown']]}).to_excel(writer, sheet_name="Dashboard", index=False)
            
            s_val = stats['wallets']['STOCK']['assets'] + stats['wallets']['STOCK']['balance']
            c_val = stats['wallets']['CRYPTO']['assets'] + stats['wallets']['CRYPTO']['balance']
            pd.DataFrame({"Danh Mục": ["STOCK", "CRYPTO"], "Tài Sản": [s_val, c_val]}).to_excel(writer, sheet_name="Dashboard", index=False, startrow=10)
            pd.DataFrame({"Danh Mục": ["STOCK", "CRYPTO"], "Tài Sản Hiện Tại": [s_val, c_val], "Lãi/Lỗ Đã Chốt": [stats['wallets']['STOCK']['realized'], stats['wallets']['CRYPTO']['realized']], "Lãi/Lỗ Đang Gồng": [stats['wallets']['STOCK']['unrealized'], stats['wallets']['CRYPTO']['unrealized']]}).to_excel(writer, sheet_name="Dashboard", index=False, startrow=15)

            port_cols = ["Ví", "Mã", "Số Lượng", "Giá Vốn TB", "Giá Hiện Tại", "Vốn Gốc VNĐ", "Lãi/Lỗ Tạm Tính"]
            portfolio = []
            for h in raw['holdings']:
                pl = (h['quantity'] * h['current_price'] * (float(raw['settings'].get('crypto_rate', 25000)) if h['wallet_id'] == 'CRYPTO' else 1)) - h['cost_basis_vnd']
                portfolio.append([h['wallet_id'], h['symbol'], h['quantity'], h['average_price'], h['current_price'], h['cost_basis_vnd'], pl])
            pd.DataFrame(portfolio, columns=port_cols).to_excel(writer, sheet_name="Portfolio", index=False)

            perf_dict = {}
            for t in raw['transactions']:
                if t['type'] == 'BAN' and t['realized_pl'] is not None and t['symbol']:
                    sym = t['symbol']
                    if sym not in perf_dict: perf_dict[sym] = {"Số Lần Bán": 0, "Tổng Lãi/Lỗ Thực Tế": 0}
                    perf_dict[sym]["Số Lần Bán"] += 1; perf_dict[sym]["Tổng Lãi/Lỗ Thực Tế"] += t['realized_pl']
            pd.DataFrame([[k, v["Số Lần Bán"], v["Tổng Lãi/Lỗ Thực Tế"]] for k, v in perf_dict.items()], columns=["Mã", "Số Lần Bán", "Tổng Lãi/Lỗ Thực Tế"]).to_excel(writer, sheet_name="Performance", index=False)

            # 2. History Sheets (startrow=16)
            l_cols = ["ID", "Loại", "Ví", "Mã", "Số Lượng", "Giá", "Thành Tiền", "Lãi Chốt", "Ghi Chú"]
            df_l = pd.DataFrame([[t['id'], t['type'], t['wallet_id'], t['symbol'], t['quantity'], t['price'], t['amount'], t['realized_pl'], t['note']] for t in raw['transactions']], columns=l_cols)
            df_l.to_excel(writer, sheet_name="Sổ Cái (All)", index=False)
            for name, wid in [("LS Nạp Rút", "CASH"), ("LS Chứng Khoán", "STOCK"), ("LS Crypto", "CRYPTO")]:
                df_sub = df_l[df_l['Ví'] == wid] if wid != "CASH" else df_l[df_l['Loại'].isin(['NAP','RUT'])]
                df_sub.to_excel(writer, sheet_name=name, index=False, startrow=16)

            # 3. Analytics
            heat_dict = {}
            for t in raw['transactions']:
                if t.get('realized_pl'):
                    dt = datetime.date.today()
                    if t.get('note'):
                        m = re.search(r'\[(\d{4}-\d{2}-\d{2})\]', str(t['note']))
                        if m: dt = datetime.datetime.strptime(m.group(1), '%Y-%m-%d').date()
                    y, mo = dt.year, dt.month
                    if y not in heat_dict: heat_dict[y] = {i: 0 for i in range(1, 13)}
                    heat_dict[y][mo] += t['realized_pl']
            pd.DataFrame([{"Năm": y} | {f"Tháng {m}": heat_dict[y][m] for m in range(1, 13)} for y in sorted(heat_dict.keys())]).to_excel(writer, sheet_name="Heatmap", index=False)

            tot = stats['total_assets']
            re_data = [{"Tài Sản": "STOCK", "Tỷ Trọng (%)": (s_val/tot*100) if tot else 0, "Mục Tiêu (%)": 40, "Lệch (VNĐ)": (0.4*tot) - s_val}, {"Tài Sản": "CRYPTO", "Tỷ Trọng (%)": (c_val/tot*100) if tot else 0, "Mục Tiêu (%)": 40, "Lệch (VNĐ)": (0.4*tot) - c_val}, {"Tài Sản": "CASH", "Tỷ Trọng (%)": (stats['wallets']['CASH']['balance']/tot*100) if tot else 0, "Mục Tiêu (%)": 20, "Lệch (VNĐ)": (0.2*tot) - stats['wallets']['CASH']['balance']}]
            pd.DataFrame(re_data).to_excel(writer, sheet_name="Rebalancing", index=False)

            w_t = [t['realized_pl'] for t in raw['transactions'] if (t.get('realized_pl') or 0) > 0]
            l_t = [t['realized_pl'] for t in raw['transactions'] if (t.get('realized_pl') or 0) < 0]
            avg_w = sum(w_t)/len(w_t) if w_t else 0; avg_l = sum(l_t)/len(l_t) if l_t else 0
            pd.DataFrame([{"Chỉ Số": "Largest Win", "Giá Trị": max(w_t) if w_t else 0}, {"Chỉ Số": "Largest Loss", "Giá Trị": min(l_t) if l_t else 0}, {"Chỉ Số": "R:R Ratio", "Giá Trị": abs(avg_w/avg_l) if avg_l else 0}]).to_excel(writer, sheet_name="Trade Analytics", index=False)

            # ==========================================
            # FORMAT EXCEL AUTO BY OPENPYXL
            # ==========================================
            wb = writer.book
            green_font = Font(color='137333', bold=True); red_font = Font(color='C5221F', bold=True)
            rule_green = CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=True, font=green_font)
            rule_red = CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, font=red_font)

            sheets_to_format = [("Portfolio", 1), ("Performance", 1), ("Sổ Cái (All)", 1), ("LS Nạp Rút", 17), ("LS Chứng Khoán", 17), ("LS Crypto", 17), ("Heatmap", 1), ("Rebalancing", 1), ("Trade Analytics", 1)]
            for name, start_row in sheets_to_format:
                ws = wb[name]
                # Dãn cột 22
                for col in range(1, ws.max_column + 1): ws.column_dimensions[get_column_letter(col)].width = 22
                
                # Chèn Summary cho các sheet lịch sử
                if name in ["LS Chứng Khoán", "LS Crypto", "LS Nạp Rút"]:
                    w_k = "STOCK" if "Chứng Khoán" in name else ("CRYPTO" if "Crypto" in name else "CASH")
                    w_d = stats['wallets'].get(w_k, {})
                    ws['A1'] = f"📑 BÁO CÁO TÀI CHÍNH: {name.upper()}"; ws['A1'].font = Font(bold=True, size=14)
                    if w_k != "CASH":
                        lines = [f"💰 Tổng giá trị: {format_currency(w_d['assets'] + w_d['balance'])}", f"💵 Tổng vốn gốc: {format_currency(w_d['total_in'] - w_d['total_out'])}", f"📈 Lãi/Lỗ: {format_currency(w_d['realized'] + w_d['unrealized'])}", f"📤 Nạp ví: {format_currency(w_d['total_in'])} | 📥 Rút ví: {format_currency(w_d['total_out'])}"]
                    else:
                        lines = [f"💰 Tổng tài sản: {format_currency(stats['total_assets'])}", f"📤 Tổng nạp: {format_currency(stats['total_in'])}", f"📥 Tổng rút: {format_currency(stats['total_out'])}", f"📈 Lãi/Lỗ: {format_currency(stats['total_pl'])}"]
                    for i, line in enumerate(lines): ws[f'A{i+3}'] = line
                
                # Table & Formatting
                last = ws.max_row
                if last >= start_row:
                    tbl_name = f"Tbl_{name.replace(' ', '_').replace('(', '').replace(')', '')}"
                    tab = Table(displayName=tbl_name[:30], ref=f"A{start_row}:{get_column_letter(ws.max_column)}{last}")
                    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True); ws.add_table(tab)
                for row in ws.iter_rows(min_row=start_row + 1):
                    for cell in row:
                        if isinstance(cell.value, (int, float)): cell.number_format = '#,##0'
                if name == "Portfolio": ws.conditional_formatting.add(f"G2:G{last}", rule_green); ws.conditional_formatting.add(f"G2:G{last}", rule_red)
                elif name in ["Sổ Cái (All)", "LS Nạp Rút", "LS Chứng Khoán", "LS Crypto"]: 
                    col_h = f"H{start_row+1}:H{last}"; ws.conditional_formatting.add(col_h, rule_green); ws.conditional_formatting.add(col_h, rule_red)

            try: wb["Dashboard"].add_image(ExcelImage(self._generate_nav_chart_io(stats, raw)), "H1")
            except: pass
        return output.getvalue()
